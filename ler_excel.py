import openpyxl

wb = openpyxl.load_workbook('Agrocomputacao_FAZU_2026_Sistema (1).xlsx')

for sheet_name in wb.sheetnames:
    print(f"\n=== {sheet_name} ===")
    ws = wb[sheet_name]
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        print(row_idx, row)
        if row_idx > 20:
            break
